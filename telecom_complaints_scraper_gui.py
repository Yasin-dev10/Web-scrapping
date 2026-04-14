import os
import sys
import time
import json
import re
import threading
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from webdriver_manager.chrome import ChromeDriverManager
    from bs4 import BeautifulSoup
    DEPS_OK = True
    DEPS_ERROR = ""
except ImportError as e:
    DEPS_OK = False
    DEPS_ERROR = str(e)


# ─── Text Cleaning ────────────────────────────────────────────────────────────
JUNK_PATTERNS = [
    re.compile(r"^\d+\s*(likes?|comments?|shares?)\s*$", re.I),
    re.compile(r"^(see more|see translation|write a comment\.*)$", re.I),
    re.compile(r"^(monday|tuesday|wednesday|thursday|friday|saturday|sunday)$", re.I),
    re.compile(r"^\d+\s*[hmd]\b$", re.I),
    re.compile(r"^(like|comment|share|reply|follow)$", re.I),
]

def is_junk(text):
    text = text.strip()
    if len(text) < 8:
        return True
    for pat in JUNK_PATTERNS:
        if pat.match(text):
            return True
    words = text.split()
    if all(w.startswith(("http", "#", "@")) for w in words):
        return True
    return False

def clean_text(text):
    text = re.sub(r"(See more|See translation)\s*", "", text, flags=re.I)
    text = re.sub(r"\n{2,}", "\n", text)
    text = re.sub(r" {2,}", " ", text)
    return text.strip()


# ─── Complaint Classifier ─────────────────────────────────────────────────────
PROBLEM_WORDS = [
    "go'ay", "go'an", "daciif", "maqan", "xun", "jaray", "goosat", "goost",
    "qaali", "dhib", "liit", "qasaaro", "tuug", "dhacay", "gaabis",
    "cabasho", "cilad", "halaysan", "xumaa", "khiyaano", "khiyaam",
    "ceeb", "eber", "baxsad", "masuuliyad darro", "isku xishooda",
    "problem", "issue", "complaint", "slow", "disconnected", "failed",
    "not working", "error", "outage", "overcharge", "stolen", "fraud",
    "bad", "worst", "terrible", "awful", "useless", "scam",
]

NETWORK_CTX   = ["shabakad", "khadka", "internet", "data", "network", "wifi",
                  "4g", "5g", "lte", "signal", "connection", "isgaarsiinta", "speed"]
BILLING_CTX   = ["lacag", "haraag", "evc", "edahab", "sahalin", "zaad", "balance",
                  "goyn", "charge", "bill", "payment", "recharge", "deduct", "money"]
SERVICE_CTX   = ["adeeg", "macaamiil", "shirkad", "somtel", "hormuud", "golis",
                  "nationlink", "telesom", "amali", "support", "customer", "service",
                  "sim", "number", "account", "activate", "block"]
POSITIVE_WORDS = ["mahadsan", "fiican", "wacan", "shidan", "macaan", "wanaag",
                   "guul", "horumar", "thanks", "great", "excellent", "good",
                   "happy", "satisfied", "perfect", "love", "best"]

def classify(text):
    t = (text or "").lower()
    prob_count  = sum(1 for p in PROBLEM_WORDS if p in t)
    has_network = any(w in t for w in NETWORK_CTX)
    has_billing = any(w in t for w in BILLING_CTX)
    has_service = any(w in t for w in SERVICE_CTX)
    has_ctx     = has_network or has_billing or has_service
    has_pos     = any(w in t for w in POSITIVE_WORDS)

    is_comp = (prob_count >= 1 and has_ctx) or prob_count >= 2

    if is_comp:
        # Suppress false positives: positive + only 1 problem word + no strong negatives
        if has_pos and prob_count < 2 and "xun" not in t and "qaali" not in t and "bad" not in t:
            return False, "none"
        if has_network:
            cat = "Network Issue"
        elif has_billing:
            cat = "Billing Issue"
        elif has_service:
            cat = "Customer Service"
        else:
            cat = "General Complaint"
        return True, cat

    return False, "none"


# ─── Excel Export ─────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1E3A5F")
COMP_FILL     = PatternFill("solid", fgColor="FFF0EE")
NOCOMP_FILL   = PatternFill("solid", fgColor="F0FFF4")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
DATA_FONT     = Font(name="Arial", size=10)
THIN_BORDER   = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
CAT_COLORS = {
    "Network Issue":      "FF6B35",
    "Billing Issue":      "F7C59F",
    "Customer Service":   "EFEFD0",
    "General Complaint":  "E8D5B7",
    "none":               "C8E6C9",
}

def save_excel(complaints, non_complaints, out_dir):
    wb = Workbook()

    # ── Sheet 1: All Data ─────────────────────────────────────────────────────
    ws_all = wb.active
    ws_all.title = "All Data"

    headers = ["#", "Text", "Label", "Category", "Source URL", "Scraped At"]
    col_widths = [5, 80, 14, 20, 40, 20]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws_all.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws_all.column_dimensions[get_column_letter(col_idx)].width = w

    ws_all.row_dimensions[1].height = 28

    all_rows = [(r, "Complaint") for r in complaints] + \
               [(r, "Non-Complaint") for r in non_complaints]

    for row_idx, (rec, label) in enumerate(all_rows, 2):
        row_data = [
            row_idx - 1,
            rec["text"],
            label,
            rec.get("complaint_type", "none"),
            rec.get("url", ""),
            rec.get("ts", ""),
        ]
        fill = COMP_FILL if label == "Complaint" else NOCOMP_FILL
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_all.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws_all.row_dimensions[row_idx].height = 45

    # Freeze header
    ws_all.freeze_panes = "A2"

    # ── Sheet 2: Complaints Only ──────────────────────────────────────────────
    ws_c = wb.create_sheet("Complaints")
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws_c.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor="8B0000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws_c.column_dimensions[get_column_letter(col_idx)].width = w
    ws_c.row_dimensions[1].height = 28
    ws_c.freeze_panes = "A2"

    for row_idx, rec in enumerate(complaints, 2):
        cat_color = CAT_COLORS.get(rec.get("complaint_type", "none"), "FFFFFF")
        fill = PatternFill("solid", fgColor=cat_color)
        row_data = [row_idx - 1, rec["text"], "Complaint",
                    rec.get("complaint_type", ""), rec.get("url", ""), rec.get("ts", "")]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_c.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws_c.row_dimensions[row_idx].height = 45

    # ── Sheet 3: Non-Complaints Only ──────────────────────────────────────────
    ws_n = wb.create_sheet("Non-Complaints")
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws_n.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor="1B5E20")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws_n.column_dimensions[get_column_letter(col_idx)].width = w
    ws_n.row_dimensions[1].height = 28
    ws_n.freeze_panes = "A2"

    for row_idx, rec in enumerate(non_complaints, 2):
        row_data = [row_idx - 1, rec["text"], "Non-Complaint",
                    "none", rec.get("url", ""), rec.get("ts", "")]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_n.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.fill = NOCOMP_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws_n.row_dimensions[row_idx].height = 45

    # ── Sheet 4: Summary ──────────────────────────────────────────────────────
    ws_s = wb.create_sheet("Summary")
    ws_s.column_dimensions["A"].width = 30
    ws_s.column_dimensions["B"].width = 15

    def sh(row, col, val, bold=False, bg=None):
        cell = ws_s.cell(row=row, column=col, value=val)
        cell.font = Font(name="Arial", size=11, bold=bold,
                         color="FFFFFF" if bg else "000000")
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center" if col == 2 else "left", vertical="center")
        return cell

    sh(1, 1, "Category", bold=True, bg="1E3A5F")
    sh(1, 2, "Count",    bold=True, bg="1E3A5F")

    cat_counts = {}
    for rec in complaints:
        cat = rec.get("complaint_type", "General Complaint")
        cat_counts[cat] = cat_counts.get(cat, 0) + 1

    r = 2
    for cat, cnt in sorted(cat_counts.items(), key=lambda x: -x[1]):
        sh(r, 1, cat)
        sh(r, 2, cnt)
        r += 1

    sh(r,   1, "Total Complaints",     bold=True)
    sh(r,   2, f"=SUM(B2:B{r-1})",    bold=True)
    sh(r+1, 1, "Total Non-Complaints", bold=True)
    sh(r+1, 2, len(non_complaints),    bold=True)
    sh(r+2, 1, "Grand Total",          bold=True, bg="1E3A5F")
    sh(r+2, 2, f"=B{r}+B{r+1}",       bold=True, bg="1E3A5F")

    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(out_dir, f"telecom_complaints_{ts}.xlsx")
    wb.save(path)
    return path


# ─── GUI Theme ────────────────────────────────────────────────────────────────
BG = "#0f1117"; SURFACE = "#1a1d27"; CARD = "#22263a"
ACCENT = "#eb6e34"; GREEN = "#22c55e"; RED = "#ef4444"
TEXT = "#e2e8f0"; MUTED = "#64748b"; BORDER = "#2d3148"
FF = "Segoe UI"


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Telecom Complaints Scraper")
        self.geometry("960x780"); self.minsize(800, 650)
        self.configure(bg=BG)
        self._scraping = False
        self._stop_flag = threading.Event()
        self._complaints = []
        self._non_complaints = []
        self._lock = threading.Lock()
        self._build_ui()
        if not DEPS_OK:
            self._log(f"⚠ Missing dependency: {DEPS_ERROR}", "err")

    # ── Build UI ──────────────────────────────────────────────────────────────
    def _build_ui(self):
        hdr = tk.Frame(self, bg=SURFACE, height=60)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        tk.Frame(hdr, bg=ACCENT, width=4).pack(side="left", fill="y")
        tk.Label(hdr, text=" 📡 Telecom Complaints Scraper",
                 font=(FF, 15, "bold"), bg=SURFACE, fg=TEXT).pack(side="left", padx=10)
        self._badge = tk.Label(hdr, text="● IDLE", font=(FF, 9, "bold"), bg=SURFACE, fg=MUTED)
        self._badge.pack(side="right", padx=20)

        main = tk.Frame(self, bg=BG)
        main.pack(fill="both", expand=True, padx=15, pady=12)

        lp = tk.Frame(main, bg=BG, width=320); lp.pack(side="left", fill="y", padx=(0, 10))
        lp.pack_propagate(False)
        rp = tk.Frame(main, bg=BG); rp.pack(side="left", fill="both", expand=True)

        self._build_settings(lp)
        self._build_stats(lp)
        self._build_log(rp)

    def _build_settings(self, p):
        c = self._card(p, "⚙  Settings")

        self._lbl(c, "Facebook Page / Group URL")
        self._url = tk.StringVar(value="https://www.facebook.com/SomaliTelecomIssues")
        self._entry(c, self._url)

        self._lbl(c, "Target count per class (complaints & non-complaints)")
        self._target = tk.IntVar(value=50)
        tk.Spinbox(c, from_=2, to=5000, textvariable=self._target,
                   bg=SURFACE, fg=TEXT, insertbackground=TEXT,
                   buttonbackground=SURFACE, relief="flat",
                   font=(FF, 10)).pack(fill="x", ipady=5, pady=(2, 10))

        self._lbl(c, "Cookies JSON Path")
        self._cookies = tk.StringVar(value="fb_cookies.json")
        cf = tk.Frame(c, bg=CARD); cf.pack(fill="x", pady=(2, 10))
        tk.Entry(cf, textvariable=self._cookies, bg=SURFACE, fg=TEXT,
                 insertbackground=TEXT, relief="flat", font=(FF, 9),
                 highlightthickness=1, highlightbackground=BORDER
                 ).pack(side="left", fill="x", expand=True, ipady=4)
        tk.Button(cf, text="…", command=self._browse,
                  bg=SURFACE, fg=TEXT, relief="flat").pack(side="left", padx=(4, 0))

        self._lbl(c, "Output Folder")
        self._outdir = tk.StringVar(value=os.path.dirname(os.path.abspath(__file__)))
        of = tk.Frame(c, bg=CARD); of.pack(fill="x", pady=(2, 12))
        tk.Entry(of, textvariable=self._outdir, bg=SURFACE, fg=TEXT,
                 insertbackground=TEXT, relief="flat", font=(FF, 9),
                 highlightthickness=1, highlightbackground=BORDER
                 ).pack(side="left", fill="x", expand=True, ipady=4)
        tk.Button(of, text="…", command=self._browse_out,
                  bg=SURFACE, fg=TEXT, relief="flat").pack(side="left", padx=(4, 0))

        bf = tk.Frame(c, bg=CARD); bf.pack(fill="x")
        self._btn_start = tk.Button(bf, text="▶  Start", bg=ACCENT, fg="white",
                                    font=(FF, 9, "bold"), relief="flat", command=self._start)
        self._btn_start.pack(side="left", fill="x", expand=True, ipady=5, padx=(0, 2))
        self._btn_stop = tk.Button(bf, text="■  Stop", bg=RED, fg="white",
                                   font=(FF, 9, "bold"), relief="flat", command=self._stop,
                                   state="disabled")
        self._btn_stop.pack(side="left", fill="x", expand=True, ipady=5, padx=(2, 0))

    def _build_stats(self, p):
        c = self._card(p, "📊  Stats")
        sf = tk.Frame(c, bg=CARD); sf.pack(fill="x", pady=5)

        def _stat(parent, color, label):
            f = tk.Frame(parent, bg=CARD); f.pack(side="left", expand=True)
            lv = tk.Label(f, text="0", font=(FF, 22, "bold"), bg=CARD, fg=color); lv.pack()
            tk.Label(f, text=label, bg=CARD, fg=MUTED, font=(FF, 8)).pack()
            return lv

        self._sv_comp   = _stat(sf, ACCENT, "Complaints")
        self._sv_nocomp = _stat(sf, GREEN,  "Non-Complaints")

        # Progress bar
        self._prog_var = tk.DoubleVar()
        style = ttk.Style(); style.theme_use("clam")
        style.configure("T.Horizontal.TProgressbar",
                        troughcolor=SURFACE, background=ACCENT,
                        darkcolor=ACCENT, lightcolor=ACCENT, bordercolor=SURFACE)
        self._prog = ttk.Progressbar(c, variable=self._prog_var, maximum=100,
                                     style="T.Horizontal.TProgressbar")
        self._prog.pack(fill="x", pady=(8, 0))
        self._prog_lbl = tk.Label(c, text="0%", bg=CARD, fg=MUTED, font=(FF, 8))
        self._prog_lbl.pack()

    def _build_log(self, p):
        c = self._card(p, "📋  Live Log", expand=True)
        self._log_widget = tk.Text(c, bg=SURFACE, fg=TEXT,
                                   font=("Consolas", 9), relief="flat",
                                   state="disabled", wrap="word", highlightthickness=0)
        self._log_widget.pack(fill="both", expand=True)
        for tag, color in [("info","#60a5fa"), ("comp", ACCENT), ("nocomp", GREEN),
                           ("err", RED), ("ok","#34d399"), ("muted", MUTED)]:
            self._log_widget.tag_configure(tag, foreground=color)
        sb = ttk.Scrollbar(self._log_widget, command=self._log_widget.yview)
        sb.pack(side="right", fill="y")
        self._log_widget.configure(yscrollcommand=sb.set)

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _card(self, parent, title, expand=False):
        outer = tk.Frame(parent, bg=BORDER, pady=1)
        outer.pack(fill="x" if not expand else "both", expand=expand, pady=(0, 10))
        inner = tk.Frame(outer, bg=CARD, padx=12, pady=10)
        inner.pack(fill="both", expand=True)
        tk.Label(inner, text=title, font=(FF, 10, "bold"), bg=CARD, fg=TEXT
                 ).pack(anchor="w", pady=(0, 8))
        return inner

    def _lbl(self, p, t):
        tk.Label(p, text=t, bg=CARD, fg=MUTED, font=(FF, 9)).pack(anchor="w")

    def _entry(self, p, var):
        tk.Entry(p, textvariable=var, bg=SURFACE, fg=TEXT,
                 insertbackground=TEXT, relief="flat", font=(FF, 10),
                 highlightthickness=1, highlightbackground=BORDER
                 ).pack(fill="x", ipady=5, pady=(2, 10))

    def _browse(self):
        f = filedialog.askopenfilename(filetypes=[("JSON", "*.json")])
        if f: self._cookies.set(f)

    def _browse_out(self):
        d = filedialog.askdirectory()
        if d: self._outdir.set(d)

    def _log(self, msg, tag=""):
        def _i():
            self._log_widget.configure(state="normal")
            ts = datetime.now().strftime("%H:%M:%S")
            self._log_widget.insert("end", f"[{ts}] ", "muted")
            self._log_widget.insert("end", msg + "\n", tag)
            self._log_widget.see("end")
            self._log_widget.configure(state="disabled")
        self.after(0, _i)

    def _update_ui(self):
        with self._lock:
            nc = len(self._complaints)
            nn = len(self._non_complaints)
        target = self._target.get()
        total  = target * 2
        done   = min(nc + nn, total)
        pct    = (done / total * 100) if total else 0
        self.after(0, lambda: [
            self._sv_comp.configure(text=str(nc)),
            self._sv_nocomp.configure(text=str(nn)),
            self._prog_var.set(pct),
            self._prog_lbl.configure(text=f"{pct:.0f}%"),
        ])

    def _set_running(self, v):
        self._scraping = v
        self._btn_start.configure(state="disabled" if v else "normal")
        self._btn_stop.configure(state="normal" if v else "disabled")
        self._badge.configure(text="● RUNNING" if v else "● IDLE",
                              fg=GREEN if v else MUTED)

    def _start(self):
        if not DEPS_OK:
            messagebox.showerror("Missing deps", DEPS_ERROR)
            return
        with self._lock:
            self._complaints.clear()
            self._non_complaints.clear()
        self._stop_flag.clear()
        self._update_ui()
        self._set_running(True)
        threading.Thread(target=self._worker, daemon=True).start()

    def _stop(self):
        self._stop_flag.set()
        self._log("Stop requested…", "err")

    # ── Core Scraping Worker ──────────────────────────────────────────────────
    def _worker(self):
        target   = self._target.get()
        url      = self._url.get().strip()
        ck_path  = self._cookies.get().strip()
        out_dir  = self._outdir.get().strip()
        driver   = None

        try:
            # ── Setup Chrome ─────────────────────────────────────────────────
            opts = Options()
            opts.add_argument("--disable-gpu")
            opts.add_argument("--disable-notifications")
            opts.add_argument("--no-sandbox")
            opts.add_argument("--disable-dev-shm-usage")
            opts.add_argument("--disable-blink-features=AutomationControlled")
            opts.add_experimental_option("excludeSwitches", ["enable-automation"])
            opts.add_experimental_option("useAutomationExtension", False)
            opts.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                              "AppleWebKit/537.36 (KHTML, like Gecko) "
                              "Chrome/124.0.0.0 Safari/537.36")
            svc    = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=svc, options=opts)
            driver.set_window_size(1366, 900)
            wait   = WebDriverWait(driver, 6)
            self._log("✅ Chrome launched", "ok")

            # ── Load cookies ─────────────────────────────────────────────────
            if os.path.exists(ck_path):
                self._log("🍪 Loading cookies…", "info")
                driver.get("https://www.facebook.com/")
                time.sleep(2)
                try:
                    with open(ck_path, encoding="utf-8") as f:
                        cookies = json.load(f)
                    for ck in cookies:
                        try:
                            driver.add_cookie({
                                "name":   ck.get("name", ""),
                                "value":  ck.get("value", ""),
                                "domain": ".facebook.com",
                                "path":   "/",
                            })
                        except Exception:
                            pass
                    self._log("✅ Cookies loaded", "ok")
                except Exception as e:
                    self._log(f"Cookie error: {e}", "err")
            else:
                self._log("⚠ Cookie file not found — may need manual login", "err")

            # ── Navigate ──────────────────────────────────────────────────────
            self._log(f"🚀 Navigating → {url}", "info")
            driver.get(url)
            time.sleep(4)

            seen        = set()
            stall_count = 0
            MAX_STALL   = 8          # ← FIXED: was 100 (300 s), now 8 (≈24 s)
            SCROLL_PAUSE = 2.5       # seconds between scrolls

            while not self._stop_flag.is_set():
                # ── Check if done ─────────────────────────────────────────────
                with self._lock:
                    nc = len(self._complaints)
                    nn = len(self._non_complaints)
                if nc >= target and nn >= target:
                    self._log("🎉 Target reached!", "ok")
                    break

                # ── Expand "See more" buttons ─────────────────────────────────
                try:
                    for btn in driver.find_elements(
                            By.XPATH,
                            "//div[@role='button'][contains(normalize-space(.),'See more')]"
                    ):
                        try: driver.execute_script("arguments[0].click();", btn)
                        except: pass
                except: pass

                # ── Parse page ───────────────────────────────────────────────
                soup       = BeautifulSoup(driver.page_source, "html.parser")
                new_count  = 0
                now_str    = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                # Grab every `dir="auto"` div (post bodies + comments)
                for block in soup.find_all("div", dir="auto"):
                    if self._stop_flag.is_set():
                        break

                    raw     = block.get_text(" ", strip=True)
                    cleaned = clean_text(raw)

                    if is_junk(cleaned):
                        continue

                    sig = cleaned[:120].lower()
                    if sig in seen:
                        continue
                    seen.add(sig)
                    new_count += 1

                    is_comp, cat = classify(cleaned)

                    with self._lock:
                        nc = len(self._complaints)
                        nn = len(self._non_complaints)

                    record = {
                        "text":           cleaned,
                        "complaint_type": cat,
                        "url":            driver.current_url,
                        "ts":             now_str,
                    }

                    if is_comp and nc < target:
                        with self._lock:
                            self._complaints.append(record)
                        self._log(f"🔴 [{cat}] {cleaned[:50]}…", "comp")
                        self._update_ui()
                    elif not is_comp and nn < target:
                        with self._lock:
                            self._non_complaints.append(record)
                        self._log(f"🟢 [Non] {cleaned[:50]}…", "nocomp")
                        self._update_ui()

                    with self._lock:
                        nc2 = len(self._complaints)
                        nn2 = len(self._non_complaints)
                    if nc2 >= target and nn2 >= target:
                        self._stop_flag.set()
                        break

                # ── Stall detection ───────────────────────────────────────────
                if new_count == 0:
                    stall_count += 1
                    self._log(f"No new content ({stall_count}/{MAX_STALL})…", "muted")
                    if stall_count >= MAX_STALL:
                        self._log("⚠ Page exhausted — stopping.", "err")
                        break
                else:
                    stall_count = 0

                # ── Scroll down ───────────────────────────────────────────────
                driver.execute_script(
                    "window.scrollTo({top: document.body.scrollHeight, behavior: 'smooth'});"
                )
                time.sleep(SCROLL_PAUSE)

            # ── Save results ──────────────────────────────────────────────────
            with self._lock:
                comps   = list(self._complaints)
                nocomps = list(self._non_complaints)

            if comps or nocomps:
                path = save_excel(comps, nocomps, out_dir)
                self._log(f"💾 Saved → {path}", "ok")
            else:
                self._log("No data collected — nothing saved.", "err")

        except Exception as ex:
            self._log(f"ERROR: {ex}", "err")
        finally:
            if driver:
                try: driver.quit()
                except: pass
            self.after(0, lambda: self._set_running(False))
            self._log("Done.", "info")


if __name__ == "__main__":
    App().mainloop()